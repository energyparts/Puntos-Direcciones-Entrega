import os, io, math, random, re
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

PTS10  = [50, 25, 20, 17, 13, 10]   # G67,G5,G4,G3,G2,G1
COST_U = [1470, 735, 588, 499.8, 382.2, 294]
EP_DARK  = "0D1F35"
EP_MID   = "1E3A5F"
EP_GREEN = "00C864"
GRP_COLS = ["4AADAD","D4A017","D96B64","5CA869","8E6BA0","D4834A"]
TAB_COLS = ["2E75B6","375623","843C0C","6A0572","7F4F24","1F4E79","0D5C63","4A1942"]


# ── Helpers ───────────────────────────────────────────────────────────────────
def cv_score(combo):
    m = sum(combo) / len(combo)
    return math.sqrt(sum((x - m)**2 for x in combo) / len(combo)) / m if m else 999


def parse_addresses_excel(file_obj):
    """Parse the pivoted Excel: col0=Etiquetas de fila, col1=Suma de Puntos"""
    df = pd.read_excel(file_obj, header=0)
    addresses = []
    for _, row in df.iterrows():
        nombre = str(row.iloc[0]).strip()
        try:
            pts = float(row.iloc[1])
        except:
            continue
        if nombre.lower() in ('total general', 'nan', ''):
            continue
        addresses.append({"nombre": nombre, "puntos": round(pts, 2)})
    return addresses


def find_combos_exhaustive(target10, max_n=50):
    """Búsqueda exhaustiva para targets pequeños (≤500)"""
    results = []; seen = set()
    for g67 in range(min(target10 // 50, 30), -1, -1):
        r1 = target10 - g67 * 50
        for g5 in range(min(r1 // 25, 30), -1, -1):
            r2 = r1 - g5 * 25
            for g4 in range(min(r2 // 20, 30), -1, -1):
                r3 = r2 - g4 * 20
                for g3 in range(min(r3 // 17, 30), -1, -1):
                    r4 = r3 - g3 * 17
                    for g2 in range(min(r4 // 13, 30), -1, -1):
                        r5 = r4 - g2 * 13
                        if r5 >= 0 and r5 % 10 == 0:
                            g1 = r5 // 10
                            k = (g67, g5, g4, g3, g2, g1)
                            if k not in seen and any(x > 0 for x in k):
                                seen.add(k); results.append(k)
                            break
    results.sort(key=cv_score)
    return results[:max_n]


def find_combos_random(target10, n=15, seed=42):
    random.seed(seed); results = []; seen = set()
    min_r = lambda f: sum(PTS10[f:])
    for _ in range(3_000_000):
        if len(results) >= n * 8: break
        mx = target10 // 50
        if mx < 1: break
        g67 = random.randint(1, min(mx, 80))
        r1  = target10 - g67 * 50
        if r1 < min_r(1): continue
        g5  = random.randint(1, min(r1 // 25, 50))
        r2  = r1 - g5 * 25
        if r2 < min_r(2): continue
        g4  = random.randint(1, min(r2 // 20, 50))
        r3  = r2 - g4 * 20
        if r3 < min_r(3): continue
        placed = False
        for dg3 in range(0, -9, -1):
            g3 = min(r3 // 17, 50) + dg3
            if g3 < 1: break
            r4 = r3 - g3 * 17
            if r4 < PTS10[4] + PTS10[5]: continue
            for dg2 in range(0, -9, -1):
                g2 = min(r4 // 13, 50) + dg2
                if g2 < 1: break
                r5 = r4 - g2 * 13
                if r5 >= 10 and r5 % 10 == 0:
                    g1 = r5 // 10
                    if g1 >= 1:
                        k = (g67, g5, g4, g3, g2, g1)
                        if k not in seen:
                            seen.add(k); results.append(k)
                        placed = True; break
            if placed: break
    results.sort(key=cv_score)
    final = []; seen2 = set()
    for c in results:
        if c not in seen2: seen2.add(c); final.append(list(c))
        if len(final) >= n: break
    return final


def generate_for_address(pts, n=15):
    target10 = round(pts * 10)
    if target10 <= 500:
        return [list(c) for c in find_combos_exhaustive(target10, max(n, 30))]
    return find_combos_random(target10, n)


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(addresses_with_combos, n_combos):
    FILL = lambda c: PatternFill("solid", start_color=c)
    THIN = Side(style='thin', color="BBBBBB")
    def brd(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hc(ws, r, c, v, bg=EP_DARK, sz=10, wt=False):
        x = ws.cell(row=r, column=c, value=v)
        x.font = Font(bold=True, name="Arial", size=sz, color="FFFFFF")
        x.fill = FILL(bg); x.border = brd()
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wt)
        return x

    def dc(ws, r, c, v, bg="FFFFFF", bold=False, fmt=None, color="000000", align="center"):
        x = ws.cell(row=r, column=c, value=v)
        x.font = Font(bold=bold, name="Arial", size=10, color=color)
        x.fill = FILL(bg); x.border = brd()
        x.alignment = Alignment(horizontal=align, vertical="center")
        if fmt: x.number_format = fmt
        return x

    wb = Workbook()

    # ── RESUMEN ───────────────────────────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Resumen"
    for i, w in enumerate([5, 32, 14, 16, 14, 16], 1):
        ws0.column_dimensions[get_column_letter(i)].width = w

    ws0.merge_cells("A1:F1")
    ws0["A1"] = "ENERGY PARTS · COMBINACIONES POR DIRECCIÓN DE ENTREGA"
    ws0["A1"].font = Font(bold=True, name="Arial", size=14, color="FFFFFF")
    ws0["A1"].fill = FILL(EP_DARK)
    ws0["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 32

    ws0.merge_cells("A2:F2")
    ws0["A2"] = "Puntos disponibles por dirección → combinaciones de grupos G1–G7 que los saldan exactamente"
    ws0["A2"].font = Font(italic=True, name="Arial", size=10, color=EP_GREEN)
    ws0["A2"].fill = FILL(EP_MID)
    ws0["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[2].height = 16

    for c, h in enumerate(["#", "Dirección de Entrega", "Puntos", "$Equivalente", "Combinaciones", "Ver hoja"], 1):
        hc(ws0, 4, c, h)

    total_pts = 0
    for i, (nombre, pts, combos) in enumerate(addresses_with_combos, 1):
        total_pts += pts
        bg = "F0F4F8" if i % 2 == 0 else "FFFFFF"
        dc(ws0, 4+i, 1, i, bg=bg, bold=True)
        dc(ws0, 4+i, 2, nombre, bg=bg, bold=True, align="left")
        dc(ws0, 4+i, 3, pts, bg=bg, bold=True, fmt='#,##0.0" pts"', color=EP_MID)
        dc(ws0, 4+i, 4, pts*294, bg=bg, fmt='"$"#,##0.00')
        dc(ws0, 4+i, 5, len(combos), bg=bg)
        dc(ws0, 4+i, 6, nombre[:20], bg=bg)
        ws0.row_dimensions[4+i].height = 16

    # Total
    rt = 4 + len(addresses_with_combos) + 1
    hc(ws0, rt, 1, "", bg=EP_MID)
    dc(ws0, rt, 2, "TOTAL", bold=True, bg=EP_MID, color="FFFFFF")
    dc(ws0, rt, 3, total_pts, bold=True, bg=EP_MID, color=EP_GREEN, fmt='#,##0.0" pts"')
    dc(ws0, rt, 4, total_pts*294, bold=True, bg=EP_MID, color="FFFFFF", fmt='"$"#,##0.00')
    dc(ws0, rt, 5, "", bg=EP_MID); dc(ws0, rt, 6, "", bg=EP_MID)
    ws0.row_dimensions[rt].height = 18

    # Ref table
    r2 = rt + 2
    ws0.merge_cells(f"A{r2}:F{r2}")
    ws0.cell(r2, 1).value = "GRUPOS DE REFERENCIA"
    ws0.cell(r2, 1).font = Font(bold=True, name="Arial", size=11, color=EP_DARK)
    ws0.cell(r2, 1).alignment = Alignment(horizontal="center")
    for c, h in enumerate(["Grupo","Puntos","Costo/u ($)","Grupo","Puntos","Costo/u ($)"], 1):
        hc(ws0, r2+1, c, h, bg=EP_MID)
    ref = [("G1",1.0,294),("G2",1.3,382.2),("G3",1.7,499.8),("G4",2.0,588),("G5",2.5,735),("G6",5.0,1470),("G7",5.0,1470)]
    for i, (g, p, co) in enumerate(ref[:4], 1):
        dc(ws0,r2+1+i,1,g,bold=True); dc(ws0,r2+1+i,2,p,fmt="0.0"); dc(ws0,r2+1+i,3,co,fmt='"$"#,##0.00')
    for i, (g, p, co) in enumerate(ref[4:], 1):
        dc(ws0,r2+1+i,4,g,bold=True); dc(ws0,r2+1+i,5,p,fmt="0.0"); dc(ws0,r2+1+i,6,co,fmt='"$"#,##0.00')

    # ── HOJA POR DIRECCIÓN ────────────────────────────────────────────────────
    for idx, (nombre, pts, combos) in enumerate(addresses_with_combos):
        safe = re.sub(r'[\\/*?:\[\]]', '_', nombre)[:31]
        ws = wb.create_sheet(safe)
        for i, w in enumerate([4,10,9,9,9,9,9,10,11,14,14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        tcol = TAB_COLS[idx % len(TAB_COLS)]

        ws.merge_cells("A1:K1")
        ws["A1"] = f"{nombre}  ·  {pts} pts  =  ${pts*294:,.2f}"
        ws["A1"].font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
        ws["A1"].fill = FILL(tcol)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:K2")
        ws["A2"] = f"Cada combinación suma exactamente {pts} pts → ${pts*294:,.2f}  |  {len(combos)} combinaciones  |  Ordenadas por uniformidad"
        ws["A2"].font = Font(italic=True, name="Arial", size=9, color=EP_GREEN)
        ws["A2"].fill = FILL(EP_MID)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

        hdrs = ["#","G6/G7\n(5pts)","G5\n(2.5)","G4\n(2)","G3\n(1.7)","G2\n(1.3)","G1\n(1)","Total\nGrupos","Total\nPuntos","Total\n$Costo","Uniformidad"]
        for c, h in enumerate(hdrs, 1):
            hc(ws, 4, c, h, bg=tcol, wt=True)
        ws.row_dimensions[4].height = 32

        for i, combo in enumerate(combos, 1):
            g67,g5,g4,g3,g2,g1 = combo
            vals = [g67,g5,g4,g3,g2,g1]
            tu  = sum(vals)
            tp  = sum(vals[k]*PTS10[k] for k in range(6))/10
            tc  = sum(vals[k]*COST_U[k] for k in range(6))
            cv  = cv_score(vals)
            badge = "★ MUY UNIFORME" if cv<0.30 else "✓ UNIFORME" if cv<0.55 else "VÁLIDA"
            bcol  = "00A550" if cv<0.30 else "1E5FAF" if cv<0.55 else "D97B00"
            bg    = "F5FAF7" if i%2==0 else "FFFFFF"

            dc(ws, 4+i, 1, i, bg=bg)
            for j, (v, gc) in enumerate(zip(vals, GRP_COLS), 2):
                x = ws.cell(row=4+i, column=j)
                x.value = v if v > 0 else "—"
                x.font  = Font(bold=v>0, name="Arial", size=10)
                x.fill  = FILL(gc+"22" if v>0 else "F8F8F8")
                x.border = brd()
                x.alignment = Alignment(horizontal="center", vertical="center")
            dc(ws, 4+i, 8, tu, bg=bg, bold=True)
            dc(ws, 4+i, 9, tp, bg=bg, bold=True, fmt='#,##0.0" pts"')
            dc(ws, 4+i, 10, tc, bg=bg, bold=True, fmt='"$"#,##0.00')
            bx = ws.cell(4+i, 11); bx.value = badge
            bx.font  = Font(bold=True, name="Arial", size=9, color=bcol)
            bx.fill  = FILL("E8F5EE" if cv<0.30 else "EEF4FF" if cv<0.55 else "FFF8EC")
            bx.border = brd()
            bx.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[4+i].height = 15

        nr = 4 + len(combos) + 2
        ws.merge_cells(f"A{nr}:K{nr}")
        ws.cell(nr, 1).value = "✔ G6 y G7 intercambiables (5 pts / $1,470 c/u)  ·  ★ muy uniforme  ✓ uniforme  · válida"
        ws.cell(nr, 1).font = Font(italic=True, name="Arial", size=9, color="777777")
        ws.cell(nr, 1).alignment = Alignment(horizontal="center")
        ws.row_dimensions[nr].height = 14

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/parse", methods=["POST"])
def parse_route():
    if "file" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    try:
        rows = parse_addresses_excel(request.files["file"])
        return jsonify({"rows": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate", methods=["POST"])
def generate_route():
    data      = request.json
    selected  = data.get("selected", [])   # [{nombre, puntos}]
    n_combos  = int(data.get("numCombos", 15))
    results   = []
    for item in selected:
        pts    = float(item["puntos"])
        combos = generate_for_address(pts, n_combos)
        results.append({"nombre": item["nombre"], "puntos": pts, "count": len(combos), "combos": combos})
    return jsonify({"results": results})


@app.route("/api/download", methods=["POST"])
def download_route():
    data     = request.json
    results  = data.get("results", [])
    n_combos = int(data.get("numCombos", 15))
    payload  = [(r["nombre"], r["puntos"], r["combos"]) for r in results]
    buf      = build_excel(payload, n_combos)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="Combinaciones_por_Direccion.xlsx")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
